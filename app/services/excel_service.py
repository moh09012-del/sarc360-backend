"""
SARC360 ERP — Excel Import/Export Service
نموذج Excel للاستيراد والعرض
"""
import io
import uuid
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_employees_template() -> bytes:
    """Generate employees.xlsx template for bulk upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Headers
    headers = [
        "Full Name (EN)",
        "Full Name (AR)",
        "Employee ID",
        "Nationality",
        "Job Title",
        "Department",
        "Iqama Number",
        "Iqama Expiry",
        "Email",
        "Phone",
        "Basic Salary (SAR)",
        "Housing Allowance (SAR)",
        "Transport Allowance (SAR)",
        "Status",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Example row
    example = [
        "Mohammed Al-Otaibi",
        "محمد العتيبي",
        "EMP-001",
        "Saudi",
        "HSE Officer",
        "operations",
        "2xxxxxxxxx",
        "2026-06-30",
        "mohammed@sarc.sa",
        "+966501234567",
        "8000.00",
        "2000.00",
        "1000.00",
        "active",
    ]

    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_clients_template() -> bytes:
    """Generate clients.xlsx template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = [
        "Client Name (EN)",
        "Client Name (AR)",
        "Commercial Registration",
        "VAT Number",
        "Billing Email",
        "Phone",
        "Address Line 1",
        "City",
        "Country",
        "Status",
    ]

    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    example = [
        "Saudi Aramco",
        "أرامكو السعودية",
        "1010001234",
        "310000000000003",
        "billing@aramco.com",
        "+966138765432",
        "Dhahran",
        "Dhahran",
        "SA",
        "active",
    ]

    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_projects_template() -> bytes:
    """Generate projects.xlsx template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    headers = [
        "Project Name (EN)",
        "Project Name (AR)",
        "Project Number",
        "Client Name",
        "Start Date",
        "End Date",
        "PO Number",
        "PO Value (SAR)",
        "Contract Value (SAR)",
        "Department",
        "Location",
        "Status",
    ]

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    example = [
        "Aramco Ras Tanura HSE",
        "أرامكو رأس تنورة السلامة",
        "PRJ-001",
        "Saudi Aramco",
        "2026-01-15",
        "2026-12-31",
        "PO-ARA-2601-001",
        "850000.00",
        "850000.00",
        "operations",
        "Ras Tanura",
        "active",
    ]

    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_timesheets_template() -> bytes:
    """Generate timesheets.xlsx template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheets"

    headers = [
        "Employee ID",
        "Project Number",
        "Week Start Date",
        "Sunday Hours",
        "Monday Hours",
        "Tuesday Hours",
        "Wednesday Hours",
        "Thursday Hours",
        "Friday Hours",
        "Saturday Hours",
        "Notes",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    example = [
        "EMP-001",
        "PRJ-001",
        "2026-03-23",
        "8",
        "8",
        "8",
        "8",
        "8",
        "0",
        "0",
        "Regular week",
    ]

    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_expenses_template() -> bytes:
    """Generate expenses.xlsx template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = [
        "Project Number",
        "Expense Date",
        "Category",
        "Description",
        "Amount Net (SAR)",
        "VAT Amount (SAR)",
        "Amount Gross (SAR)",
        "Status",
    ]

    header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    example = [
        "PRJ-001",
        "2026-03-20",
        "materials",
        "Safety equipment",
        "5000.00",
        "750.00",
        "5750.00",
        "approved",
    ]

    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
