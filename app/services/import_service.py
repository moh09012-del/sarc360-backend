"""
SARC360 ERP — Excel Data Import Service
خدمة استيراد بيانات Excel
"""
import uuid
from datetime import datetime, date
from decimal import Decimal
from typing import Any, Dict, List

from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee, EmploymentType, EmployeeStatus, Department
from app.models.client import Client
from app.models.project import Project
from app.models.timesheet import Timesheet
from app.models.expense import Expense
from app.schemas.employee import EmployeeCreate
from app.crud.employee import employee_crud


class ImportResult:
    """Result of bulk import operation."""

    def __init__(self):
        self.total_rows = 0
        self.success_rows = 0
        self.failed_rows = 0
        self.errors: List[Dict[str, Any]] = []

    def add_error(self, row_num: int, error: str, values: Dict = None):
        self.failed_rows += 1
        self.errors.append(
            {
                "row": row_num,
                "error": error,
                "values": values or {},
            }
        )

    def to_dict(self):
        return {
            "total_rows": self.total_rows,
            "success_rows": self.success_rows,
            "failed_rows": self.failed_rows,
            "errors": self.errors,
        }


async def import_employees_excel(
    db: AsyncSession,
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    created_by: uuid.UUID,
) -> ImportResult:
    """Import employees from Excel file."""
    result = ImportResult()

    try:
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        result.add_error(0, f"Invalid Excel file: {str(e)}")
        return result

    # Skip header row, start from row 2
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result.total_rows += 1

        try:
            # Parse row
            (
                full_name_en,
                full_name_ar,
                emp_id,
                nationality,
                job_title,
                department,
                iqama_number,
                iqama_expiry_str,
                email,
                phone,
                basic_salary_str,
                housing_str,
                transport_str,
                status,
            ) = row[:14]

            # Validate required fields
            if not full_name_en or not nationality or not job_title:
                result.add_error(row_idx, "Missing required fields: Full Name, Nationality, Job Title")
                continue

            # Parse salary values
            try:
                basic_salary = Decimal(str(basic_salary_str or 0))
                housing_allowance = Decimal(str(housing_str or 0))
                transport_allowance = Decimal(str(transport_str or 0))
            except Exception as e:
                result.add_error(
                    row_idx,
                    f"Invalid salary format: {str(e)}",
                    {"row": row},
                )
                continue

            # Parse date
            iqama_expiry = None
            if iqama_expiry_str:
                try:
                    if isinstance(iqama_expiry_str, str):
                        iqama_expiry = datetime.strptime(iqama_expiry_str, "%Y-%m-%d").date()
                    else:
                        iqama_expiry = iqama_expiry_str  # Already a date object
                except Exception:
                    result.add_error(row_idx, "Invalid date format for Iqama Expiry (use YYYY-MM-DD)")
                    continue

            # Create employee
            emp_create = EmployeeCreate(
                full_name_en=full_name_en,
                full_name_ar=full_name_ar,
                nationality=nationality,
                job_title=job_title,
                department=department or "operations",
                iqama_number=iqama_number,
                iqama_expiry_date=iqama_expiry,
                basic_salary_sar=basic_salary,
                housing_allowance_sar=housing_allowance,
                transport_allowance_sar=transport_allowance,
                status=status or "active",
                tenant_id=tenant_id,
            )

            # Create in DB
            emp = await employee_crud.create(db, emp_create, created_by=created_by)
            result.success_rows += 1
            await db.commit()

        except Exception as e:
            result.add_error(row_idx, f"Error creating employee: {str(e)}", {"row": row})
            continue

    return result


async def import_clients_excel(
    db: AsyncSession,
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    created_by: uuid.UUID,
) -> ImportResult:
    """Import clients from Excel file."""
    result = ImportResult()

    try:
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        result.add_error(0, f"Invalid Excel file: {str(e)}")
        return result

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result.total_rows += 1

        try:
            (
                name_en,
                name_ar,
                cr_number,
                vat_number,
                billing_email,
                phone,
                address_line1,
                city,
                country,
                is_active_str,
            ) = row[:10]

            if not name_en:
                result.add_error(row_idx, "Missing client name")
                continue

            is_active = is_active_str != "inactive" if is_active_str else True

            client = Client(
                tenant_id=tenant_id,
                name_en=name_en,
                name_ar=name_ar,
                cr_number=cr_number,
                vat_number=vat_number,
                billing_email=billing_email,
                phone=phone,
                address_line1=address_line1,
                city=city,
                country=country or "SA",
                is_active=is_active,
            )
            db.add(client)
            result.success_rows += 1

        except Exception as e:
            result.add_error(row_idx, f"Error creating client: {str(e)}", {"row": row})
            continue

    await db.commit()
    return result


async def import_projects_excel(
    db: AsyncSession,
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    created_by: uuid.UUID,
) -> ImportResult:
    """Import projects from Excel file."""
    result = ImportResult()

    try:
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        result.add_error(0, f"Invalid Excel file: {str(e)}")
        return result

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result.total_rows += 1

        try:
            (
                name_en,
                name_ar,
                proj_number,
                client_name,
                start_date_str,
                end_date_str,
                po_number,
                po_value_str,
                contract_value_str,
                department,
                location,
                status,
            ) = row[:12]

            if not name_en or not client_name or not proj_number:
                result.add_error(row_idx, "Missing required fields: Project Name, Client, Project Number")
                continue

            # Parse dates
            try:
                start_date = (
                    datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    if isinstance(start_date_str, str)
                    else start_date_str
                )
                end_date = None
                if end_date_str:
                    end_date = (
                        datetime.strptime(end_date_str, "%Y-%m-%d").date()
                        if isinstance(end_date_str, str)
                        else end_date_str
                    )
            except Exception:
                result.add_error(row_idx, "Invalid date format (use YYYY-MM-DD)")
                continue

            # Parse values
            try:
                po_value = Decimal(str(po_value_str or 0))
                contract_value = Decimal(str(contract_value_str or 0))
            except Exception:
                result.add_error(row_idx, "Invalid amount format")
                continue

            project = Project(
                tenant_id=tenant_id,
                name_en=name_en,
                name_ar=name_ar,
                project_number=proj_number,
                client_name=client_name,
                po_number=po_number,
                po_value_sar=po_value,
                contract_value_sar=contract_value,
                start_date=start_date,
                end_date=end_date,
                department=department,
                location=location,
                status=status or "active",
                created_by=created_by,
            )
            db.add(project)
            result.success_rows += 1

        except Exception as e:
            result.add_error(row_idx, f"Error creating project: {str(e)}", {"row": row})
            continue

    await db.commit()
    return result


async def import_timesheets_excel(
    db: AsyncSession,
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    created_by: uuid.UUID,
) -> ImportResult:
    """Import timesheets from Excel file."""
    result = ImportResult()

    try:
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        result.add_error(0, f"Invalid Excel file: {str(e)}")
        return result

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result.total_rows += 1

        try:
            (
                emp_id,
                proj_number,
                week_start_str,
                sun_hrs,
                mon_hrs,
                tue_hrs,
                wed_hrs,
                thu_hrs,
                fri_hrs,
                sat_hrs,
                notes,
            ) = row[:11]

            if not emp_id or not proj_number or not week_start_str:
                result.add_error(row_idx, "Missing required fields: Employee ID, Project Number, Week Start Date")
                continue

            # Find employee and project
            emp_result = await db.execute(
                select(Employee).where(
                    Employee.employee_number == emp_id,
                    Employee.tenant_id == tenant_id,
                )
            )
            emp = emp_result.scalar_one_or_none()
            if not emp:
                result.add_error(row_idx, f"Employee '{emp_id}' not found")
                continue

            proj_result = await db.execute(
                select(Project).where(
                    Project.project_number == proj_number,
                    Project.tenant_id == tenant_id,
                )
            )
            proj = proj_result.scalar_one_or_none()
            if not proj:
                result.add_error(row_idx, f"Project '{proj_number}' not found")
                continue

            # Parse date
            try:
                week_start = (
                    datetime.strptime(week_start_str, "%Y-%m-%d").date()
                    if isinstance(week_start_str, str)
                    else week_start_str
                )
            except Exception:
                result.add_error(row_idx, "Invalid date format (use YYYY-MM-DD)")
                continue

            # Parse hours
            try:
                hours = [
                    Decimal(str(h or 0))
                    for h in [sun_hrs, mon_hrs, tue_hrs, wed_hrs, thu_hrs, fri_hrs, sat_hrs]
                ]
            except Exception:
                result.add_error(row_idx, "Invalid hours format")
                continue

            # Generate timesheet number
            ts_count_result = await db.execute(
                select(func.count()).select_from(Timesheet).where(Timesheet.tenant_id == tenant_id)
            )
            ts_count = ts_count_result.scalar_one() + 1
            ts_number = f"TS-{week_start.strftime('%Y%m')}-{ts_count:03d}"

            ts = Timesheet(
                tenant_id=tenant_id,
                timesheet_number=ts_number,
                employee_id=emp.id,
                project_id=proj.id,
                week_start_date=week_start,
                hours_sun=hours[0],
                hours_mon=hours[1],
                hours_tue=hours[2],
                hours_wed=hours[3],
                hours_thu=hours[4],
                hours_fri=hours[5],
                hours_sat=hours[6],
                notes=notes,
                status="draft",
                created_by=created_by,
            )
            db.add(ts)
            result.success_rows += 1

        except Exception as e:
            result.add_error(row_idx, f"Error creating timesheet: {str(e)}", {"row": row})
            continue

    await db.commit()
    return result


async def import_expenses_excel(
    db: AsyncSession,
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    created_by: uuid.UUID,
) -> ImportResult:
    """Import expenses from Excel file."""
    result = ImportResult()

    try:
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.active
    except Exception as e:
        result.add_error(0, f"Invalid Excel file: {str(e)}")
        return result

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result.total_rows += 1

        try:
            (
                proj_number,
                expense_date_str,
                category,
                description,
                amount_net_str,
                vat_amount_str,
                amount_gross_str,
                status,
            ) = row[:8]

            if not proj_number or not expense_date_str:
                result.add_error(row_idx, "Missing required fields: Project Number, Expense Date")
                continue

            # Find project
            proj_result = await db.execute(
                select(Project).where(
                    Project.project_number == proj_number,
                    Project.tenant_id == tenant_id,
                )
            )
            proj = proj_result.scalar_one_or_none()
            if not proj:
                result.add_error(row_idx, f"Project '{proj_number}' not found")
                continue

            # Parse date
            try:
                expense_date = (
                    datetime.strptime(expense_date_str, "%Y-%m-%d").date()
                    if isinstance(expense_date_str, str)
                    else expense_date_str
                )
            except Exception:
                result.add_error(row_idx, "Invalid date format (use YYYY-MM-DD)")
                continue

            # Parse amounts
            try:
                amount_net = Decimal(str(amount_net_str or 0))
                vat_amount = Decimal(str(vat_amount_str or 0))
                amount_gross = Decimal(str(amount_gross_str or 0))
            except Exception:
                result.add_error(row_idx, "Invalid amount format")
                continue

            expense = Expense(
                tenant_id=tenant_id,
                project_id=proj.id,
                expense_date=expense_date,
                category=category or "other",
                description=description,
                amount_net=amount_net,
                vat_amount=vat_amount,
                amount_gross=amount_gross,
                status=status or "approved",
                created_by=created_by,
            )
            db.add(expense)
            result.success_rows += 1

        except Exception as e:
            result.add_error(row_idx, f"Error creating expense: {str(e)}", {"row": row})
            continue

    await db.commit()
    return result
